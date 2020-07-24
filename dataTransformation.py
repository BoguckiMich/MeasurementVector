import pandas as pd
import glob
import openpyxl
import csv

# filenames
excel_names = glob.glob("*.xlsx")

# read them in
excels = [pd.ExcelFile(name) for name in excel_names]

# turn them into dataframes
frames = [x.parse(x.sheet_names[0], header=None, index_col=None) for x in excels]

# delete the first row for all frames except the first
# i.e. remove the header row -- assumes it's the first
# frames[1:] = [df[1:] for df in frames[1:]]

# concatenate them..
combined = pd.concat(frames)

# write it out
combined.to_excel("c.xlsx", header='1', index=False)


wb = openpyxl.load_workbook('c.xlsx')
sheet = wb.active
invert = wb.create_sheet("inverted")
# invert2 = wb.create_sheet("inverted2")


invertRow = 1
invertCol = 1

for oldRow in range (1, sheet.max_row):
	if ((sheet.cell(row=oldRow, column=1).value == sheet.cell(row=oldRow + 1, column=1).value)):
		invert.cell(row=invertRow, column=invertCol).value = sheet.cell(row=oldRow+1, column=3).value
		invertCol = invertCol + 1
	else :
		invertRow = invertRow + 1
		invertCol = 1
		invert.cell(row=invertRow, column=invertCol).value = sheet.cell(row=oldRow+1, column=1).value
		invert.cell(row=invertRow, column=invertCol+1).value = sheet.cell(row=oldRow+1, column=2).value
		invert.cell(row=invertRow, column=invertCol+2).value = sheet.cell(row=oldRow+1, column=3).value
		invertCol = 4

wb.save('10.xlsx')

with open('plot.csv', 'w', newline='', encoding="utf-8") as f:
	c = csv.writer(f)
	for r in invert.rows:
		c.writerow([cell.value for cell in r])

