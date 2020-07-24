import openpyxl
import pandas as pd
import math

#dataFrame1, dataFrame2

def palowanie_do_excela(projekt, inwentaryzacja):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pale"
    ws.sheet_properties.tabColor = "00ff00"
    ws['B1'] = 'TABELA zbiorcza'
    ws['A2'] = 'pomiar powykonawczy oczepow'
    ws['A3'] = 'Projektowana S7 – na odcinku Koszwały Kazimierzowo'
    ws['A4'] = 'Projekt'
    ws['G4'] = 'Inwentaryzacja'
    ws.merge_cells('A4:F4')
    ws.merge_cells('G4:M4')
    ws['A5'] = 'L.p.'
    ws['B5'] = 'NR \n platformy'
    ws['C5'] = 'Nr platfomry x nr pala'
    ws['D5'] = 'X \n proj'
    ws['E5'] = 'Y \n proj'
    ws['F5'] = 'H \nproj \ngory oczepu'
    ws['G5'] = 'X'
    ws['H5'] = 'Y'
    ws['I5'] = 'H\nwykonane\ngory oczepu'
    ws['J5'] = 'dX\nwyk\ncm'
    ws['K5'] = 'dY\nwyk\ncm'
    ws['L5'] = 'wektor przesunięcia Pala\ncm'
    ws['M5'] = 'dz\nwyk\ncm'

    for p in projekt.index:
        ws['A' + str(6+p)] = p+1
        ws['B' + str(6+p)] = projekt['nr'][p][:2]
        ws['C' + str(6+p)] = projekt['nr'][p]
        ws['D' + str(6+p)] = projekt['x'][p]
        ws['E' + str(6+p)] = projekt['y'][p]
        ws['F' + str(6+p)] = projekt['h'][p]
        ws['G' + str(6+p)] = inwentaryzacja['x1'][p]
        ws['H' + str(6+p)] = inwentaryzacja['y1'][p]
        ws['I' + str(6+p)] = inwentaryzacja['h'][p]
        ws['J' + str(6+p)] = (projekt['x'][p] - inwentaryzacja['x1'][p])*100
        ws['K' + str(6+p)] = (projekt['y'][p] - inwentaryzacja['y1'][p])*100
        ws['L' + str(6+p)] = round(math.sqrt((projekt['x'][p] - inwentaryzacja['x1'][p])**2 + (projekt['y'][p] - inwentaryzacja['y1'][p])**2)*100, 0)
        ws['M' + str(6+p)] = (projekt['h'][p] - inwentaryzacja['h'][p])*100

    
    wb.save("./result/palowanie.xlsx")
